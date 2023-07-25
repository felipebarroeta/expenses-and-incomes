import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
import tkinter as tk
from tkinter import ttk
import matplotlib.pyplot as plt

def save_expenses_and_income_to_excel(data):
    df = pd.DataFrame(data)

    # Calculate the savings
    df['Savings'] = df['Income'] - df['Expenses']
    df['Accumulated Savings'] = df['Savings'].cumsum()

    # Create an Excel writer object
    writer = pd.ExcelWriter('expenses_and_income.xlsx', engine='openpyxl')

    # Convert the DataFrame to an Excel sheet
    df.to_excel(writer, sheet_name='Expenses and Income', index=False)
    worksheet = writer.sheets['Expenses and Income']

    # Format the 'Savings' column with colors
    for cell in worksheet['D']:
        if cell.row == 1:
            continue  # Skip the header row
        if cell.value >= 0:
            cell.fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")  # Green color for positive savings
        else:
            cell.fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")  # Red color for negative savings

    # Save the changes and close the Excel writer
    writer._save()  # Use writer._save() if using pandas version 1.3.0 or later
    writer.close()

def create_graph(data):
    df = pd.DataFrame(data)
    df['Savings'] = df['Income'] - df['Expenses']
    df['Accumulated Savings'] = df['Savings'].cumsum()

    plt.plot(df['Month'], df['Savings'], marker='o', label='Savings')
    plt.plot(df['Month'], df['Accumulated Savings'], marker='o', label='Accumulated Savings')
    plt.xlabel('Month')
    plt.ylabel('Amount')
    plt.title('Savings and Accumulated Savings')
    plt.legend()
    plt.xticks(rotation=45)
    plt.tight_layout()
    plt.show()

def submit_data():
    data = {
        'Month': [month_entry.get() for month_entry in month_entries], 
        'Income': [int(income_entry.get()) for income_entry in income_entries],
        'Expenses': [int(expense_entry.get()) for expense_entry in expense_entries]
    }

    save_expenses_and_income_to_excel(data)
    create_graph(data)

if __name__ == "__main__":
    root = tk.Tk()
    root.title("Expenses and Income Tracker")

    label_frame = ttk.LabelFrame(root, text="Enter Expenses and Income Data")
    label_frame.grid(row=0, column=0, padx=10, pady=10)

    month_labels = ['January', 'February', 'March', 'April', 'May', 'June', 'July', 'August', 'September', 'October', 'November', 'December']

    month_entries = []
    income_entries = []
    expense_entries = []

    for i, month in enumerate(month_labels):
        ttk.Label(label_frame, text=month).grid(row=i, column=0)
        month_entry = ttk.Entry(label_frame)
        month_entry.grid(row=i, column=1)
        month_entries.append(month_entry)

        ttk.Label(label_frame, text="Income").grid(row=i, column=2)
        income_entry = ttk.Entry(label_frame)
        income_entry.grid(row=i, column=3)
        income_entries.append(income_entry)

        ttk.Label(label_frame, text="Expenses").grid(row=i, column=4)
        expense_entry = ttk.Entry(label_frame)
        expense_entry.grid(row=i, column=5)
        expense_entries.append(expense_entry)

    ttk.Button(root, text="Submit", command=submit_data).grid(row=len(month_labels)+1, column=0, columnspan=6, padx=10, pady=10)

    root.mainloop()
